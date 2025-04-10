import datetime
import os
import io
import subprocess
from datetime import timezone

import pytz
from django.conf import settings
from django.http import HttpResponse, FileResponse
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from reportlab.lib.colors import HexColor
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import landscape, A4
from pdfrw import PdfReader, PdfWriter, PageMerge
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from openpyxl.styles import Font, Alignment, Border, Side

from accounts.models import User
from main.models import Subject
from progress.models import UserSubject
from openpyxl.drawing.image import Image as XLImage


# Generate certificate
# ----------------------------------------------------------------------------------------------------------------------
@login_required
def generate_certificate_view(request, subject_pk):
    user = request.user
    subject = get_object_or_404(Subject, pk=subject_pk)
    user_subject = UserSubject.objects.filter(user=request.user, subject=subject).order_by('id').first()

    if request.method == 'POST':
        packet = io.BytesIO()

        font_path = os.path.join(settings.BASE_DIR, 'templates', 'static', 'fonts', 'Georgia', 'GeorgiaPro-Bold.ttf')
        pdfmetrics.registerFont(TTFont('GeorgiaPro', font_path))

        page_width, page_height = landscape(A4)
        can = canvas.Canvas(packet, pagesize=(page_width, page_height))

        # Full name
        # --------------------------------------------------------------------------------------------------------------
        can.setFont('GeorgiaPro', 36)
        can.setFillColor(HexColor("#a17b41"))

        full_name = f'{user.first_name} {user.last_name}'
        can.drawCentredString(page_width / 2, 340, full_name)


        # Course name
        # --------------------------------------------------------------------------------------------------------------
        course_name = f'{subject.title}'
        can.setFont('GeorgiaPro', 24)
        can.setFillColor(HexColor("#0d491d"))
        can.drawCentredString(page_width / 2, 280, course_name)

        # Course number
        # --------------------------------------------------------------------------------------------------------------
        course_name = f'N00{user_subject.id}'
        can.setFont('GeorgiaPro', 18)
        can.setFillColor(HexColor("#000000"))
        can.drawCentredString(224, 116, course_name)

        # Course completed date
        # --------------------------------------------------------------------------------------------------------------
        completed_at = user_subject.completed_at.strftime('%d.%m.%Y')
        course_name = f'{completed_at}'
        can.setFont('GeorgiaPro', 14)
        can.setFillColor(HexColor("#000000"))
        can.drawCentredString(492, 213, course_name)

        # Save all
        can.save()
        packet.seek(0)

        cert_path = os.path.join(settings.BASE_DIR, 'templates', 'static', 'images', 'cert.pdf')
        template_pdf = PdfReader(cert_path)
        overlay_pdf = PdfReader(packet)

        for page, overlay in zip(template_pdf.pages, overlay_pdf.pages):
            merger = PageMerge(page)
            merger.add(overlay).render()

        output = io.BytesIO()
        PdfWriter(output, trailer=template_pdf).write()
        output.seek(0)
        return HttpResponse(output, content_type='application/pdf')

    return HttpResponse("GET сұраныс рұқсат етілмейді.", status=405)


# Export excel user_subject
# ----------------------------------------------------------------------------------------------------------------------
def export_user_course_report_excel(request, user_subject_id):
    user_subject = UserSubject.objects.select_related('user', 'subject').get(id=user_subject_id)

    wb = Workbook()
    ws = wb.active
    ws.title = f'{user_subject.user.first_name} {user_subject.user.last_name}'

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # ===== HEADER =====
    ws.merge_cells('A2:G2')
    ws['A2'] = '“Өзбекәлі Жәнібеков атындағы Оңтүстік Қазақстан педагогикалық университеті”'
    ws['A2'].font = Font(name='Times New Roman', size=12, bold=True)
    ws['A2'].alignment = Alignment(horizontal='center', wrap_text=True)

    ws.merge_cells('A3:G3')
    ws['A3'] = 'Некоммерческое акционерное общество "Южно-Казахстанский педагогический университет имени Өзбекали Жанибекова"'
    ws['A3'].font = Font(name='Times New Roman', size=11, bold=True)
    ws['A3'].alignment = Alignment(horizontal='center', wrap_text=True)

    # ===== LOGO ORTADA =====
    logo_path = os.path.join('templates', 'static', 'images', 'zh-logo.png')  # Өз жолыңмен ауыстыр
    if os.path.exists(logo_path):
        logo = XLImage(logo_path)
        logo.width = 200
        logo.height = 80
        logo.anchor = 'C4'  # Ортасына
        ws.add_image(logo)

    # ===== ВЕДОМОСТЬ ТАҚЫРЫБЫ =====
    ws.merge_cells('A9:G9')
    ws['A9'] = 'ЖЕКЕ БІЛІМАЛУШЫ ВЕДОМОСІ'
    ws['A9'].font = Font(size=14, bold=True)
    ws['A9'].alignment = Alignment(horizontal='center')

    # ===== STUDENT INFO =====
    start_time = user_subject.created_at
    kazakhstan_timezone = pytz.timezone('Asia/Almaty')
    if start_time:
        start_time = start_time.astimezone(kazakhstan_timezone).replace(tzinfo=None)
        start_time = start_time.strftime('%Y-%m-%d %H:%M:%S')

    info = {
        'Аты-жөні:': f'{user_subject.user.last_name} {user_subject.user.first_name}',
        'Факультет:': user_subject.user.faculty,
        'Мамандық': user_subject.user.profession,
        'Топ:': user_subject.user.group,
        'Пән:': f'{user_subject.subject.title}',
        'Сабақтар саны:': f'{user_subject.user_lessons.count()}',
        'Аяқталған сабақтар саны:': user_subject.user_lessons.filter(completed=True).count(),
        'Статус:': 'Аяқталды' if user_subject.completed else 'Аяқталмады',
        'Басталған уақыты': start_time,
        'Жалпы бағасы:': f'{user_subject.total_percent}%',
    }

    row_idx = 11
    for label, value in info.items():
        ws[f'A{row_idx}'] = label
        ws[f'B{row_idx}'] = value
        row_idx += 1

    # ===== TABLE HEADER =====
    table_row = row_idx + 2
    headers = ['Сабақ', 'Сабақ уақыты', 'Тапсырма бағасы', 'Тестілеу бағасы', 'Жалпы бағалау', 'Статус', 'Орындалған уақыты']

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=table_row, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # ===== TABLE CONTENT =====
    for i, user_lesson in enumerate(user_subject.user_lessons.all(), 1):
        title = user_lesson.lesson.title
        duration = user_lesson.lesson.duration
        task_score = f'{sum([t.grade for t in user_lesson.user_tasks.all()])}%'
        test_score = f'{sum([t.score for t in user_lesson.user_tests.all()])}%' if sum([t.score for t in user_lesson.user_tests.all()]) else '-'
        lesson_score = f'{user_lesson.lesson_score}%'
        status = 'Орындалды' if user_lesson.completed else '-'
        submitted_at = (
            user_lesson.completed_at.astimezone(kazakhstan_timezone).replace(tzinfo=None)
            if user_lesson.completed_at else '-'
        )

        values = [title, duration, task_score, test_score, lesson_score, status, submitted_at]

        for j, val in enumerate(values, 1):
            cell = ws.cell(row=table_row + i, column=j, value=val)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border

    # ===== FOOTER =====
    footer_row = table_row + i + 3
    ws[f'A{footer_row}'] = 'Мұғалімнің қолы: _______________________'

    # ===== AUTO WIDTH (MAX 42) =====
    for col_idx in range(1, 1 + len(headers)):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for row in range(table_row, table_row + i + 1):
            val = ws.cell(row=row, column=col_idx).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 42)

    # ===== GRID OFF =====
    ws.sheet_view.showGridLines = False

    # ===== HIDE UNUSED COLUMNS & ROWS =====
    for col_idx in range(8, 100):
        ws.column_dimensions[get_column_letter(col_idx)].hidden = True
    for row in range(footer_row + 2, 100):
        ws.row_dimensions[row].hidden = True

    # ===== VISIBLE AREA =====
    ws.sheet_view.selection[0].sqref = f"A1:G{footer_row+1}"
    ws.print_area = f"A1:G{footer_row+1}"

    # ===== EXPORT =====
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'user_subject_{user_subject.id}_report.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# Export excel user_subject
# ----------------------------------------------------------------------------------------------------------------------
def export_user_courses_report_excel(request):
    user_subjects = UserSubject.objects.all()

    wb = Workbook()
    ws = wb.active
    ws.title = f'Жалпы топ ведомосты'

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # ===== HEADER =====
    ws.merge_cells('A2:G2')
    ws['A2'] = '“Өзбекәлі Жәнібеков атындағы Оңтүстік Қазақстан педагогикалық университеті”'
    ws['A2'].font = Font(name='Times New Roman', size=12, bold=True)
    ws['A2'].alignment = Alignment(horizontal='center', wrap_text=True)

    ws.merge_cells('A3:G3')
    ws['A3'] = 'Некоммерческое акционерное общество "Южно-Казахстанский педагогический университет имени Өзбекали Жанибекова"'
    ws['A3'].font = Font(name='Times New Roman', size=11, bold=True)
    ws['A3'].alignment = Alignment(horizontal='center', wrap_text=True)

    # ===== LOGO ORTADA =====
    logo_path = os.path.join('templates', 'static', 'images', 'zh-logo.png')  # Өз жолыңмен ауыстыр
    if os.path.exists(logo_path):
        logo = XLImage(logo_path)
        logo.width = 200
        logo.height = 80
        logo.anchor = 'C4'  # Ортасына
        ws.add_image(logo)

    # ===== ВЕДОМОСТЬ ТАҚЫРЫБЫ =====
    ws.merge_cells('A9:G9')
    ws['A9'] = 'ЖАЛПЫ ТОПТЫҢ ВЕДОМОСІ'
    ws['A9'].font = Font(size=14, bold=True)
    ws['A9'].alignment = Alignment(horizontal='center')

    # ===== STUDENT INFO =====
    group_date = datetime.datetime.now()
    info = {
        'Оқытушының аты-жөні:': f'Меруерт Бағланова Сыдыковна',
        'Білім алушылар саны': User.objects.all().count(),
        'Орындалған пәндер саны:': user_subjects.count(),
        'Жалпы пәндер саны:': f'{user_subjects.count()}',
        'Топ ведомость шығарылған уақыты': group_date.strftime('%d.%m.%Y %H:%M:%S')
    }

    row_idx = 11
    for label, value in info.items():
        ws[f'A{row_idx}'] = label
        ws[f'B{row_idx}'] = value
        row_idx += 1

    # ===== TABLE HEADER =====
    table_row = row_idx + 2
    headers = ['Пән', 'Білім алушы', 'Факультет', 'Мамандық', 'Тобы', 'Пән бағасы', 'Статус', 'Орындалған уақыты']

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=table_row, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # ===== TABLE CONTENT =====
    for i, user_subject in enumerate(user_subjects, 1):
        title = user_subject.subject.title
        full_name = f'{user_subject.user.first_name} {user_subject.user.last_name}'
        faculty = user_subject.user.faculty
        profession = user_subject.user.profession
        group = user_subject.user.group
        subject_score = f'{user_subject.total_percent}%'
        status = 'Орындалды' if user_subject.completed else '-'
        completed_at = (
            user_subject.completed_at.strftime('%d.%m.%Y %H:%M:%S')
            if user_subject.completed_at else '-'
        )

        values = [title, full_name, faculty, profession, group, subject_score, status, completed_at]

        for j, val in enumerate(values, 1):
            cell = ws.cell(row=table_row + i, column=j, value=val)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border

    # ===== FOOTER =====
    footer_row = table_row + i + 3
    ws[f'A{footer_row}'] = 'Мұғалімнің қолы: _______________________'

    # ===== AUTO WIDTH (MAX 42) =====
    for col_idx in range(1, 1 + len(headers)):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for row in range(table_row, table_row + i + 1):
            val = ws.cell(row=row, column=col_idx).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 42)

    # ===== GRID OFF =====
    ws.sheet_view.showGridLines = False

    # ===== HIDE UNUSED COLUMNS & ROWS =====
    for col_idx in range(8, 100):
        ws.column_dimensions[get_column_letter(col_idx)].hidden = True
    for row in range(footer_row + 2, 100):
        ws.row_dimensions[row].hidden = True

    # ===== VISIBLE AREA =====
    ws.sheet_view.selection[0].sqref = f"A1:G{footer_row+1}"
    ws.print_area = f"A1:G{footer_row+1}"

    # ===== EXPORT =====
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'user_subjects_report.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
